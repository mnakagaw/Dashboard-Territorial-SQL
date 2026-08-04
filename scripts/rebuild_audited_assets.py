"""Rebuild audited dashboard assets from the checked-in official source files.

This script intentionally fails when a source territory cannot be mapped to the
158-municipality index.  It updates both runtime copies (public/data) and the
source mirror (src/data) so the SQLite package and Git checkout stay aligned.
"""

from __future__ import annotations

import csv
import json
import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PUBLIC_DATA = ROOT / "public" / "data"
SOURCE_DATA = ROOT / "src" / "data"
CENSUS_SOURCE = ROOT / "data_sources" / "Datos_Originales_Censo2022"
HEALTH_SOURCE = ROOT / "data_sources" / "Establecimientos de Salud, SNS, 1878-2025.csv"
FINAL_POPULATION_SOURCE = (
    ROOT / "data_sources" / "official" / "final_population_2022_municipalities.json"
)

LEVEL_KEYS = {
    "NINGUNO": "ninguno",
    "PREPRIMARIA": "preprimaria",
    "PRIMARIA O BASICA": "primaria",
    "SECUNDARIA O MEDIA": "secundaria",
    "UNIVERSITARIA O SUPERIOR": "superior",
}

NAME_ALIASES = {
    "BAHORUCO": "BAORUCO",
    "CAMBITA GARAVITOS": "CAMBITA GARABITOS",
    "LA MATA": "VILLA LA MATA",
    # Source-table typo in C5 (Internet): the municipality is printed as
    # "de Santiago" underneath the province Santiago.
    "DE SANTIAGO": "SANTIAGO",
}


def normalize_name(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.upper().replace("Ñ", "N")
    text = re.sub(r"\s*\(D\.?\s*M\.?\)\s*", " ", text)
    text = re.sub(r"[^A-Z0-9]+", " ", text).strip()
    return NAME_ALIASES.get(text, text)


def json_load(path: Path):
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def write_asset(filename: str, value) -> None:
    payload = json.dumps(value, ensure_ascii=False, indent=2) + "\n"
    for directory in (PUBLIC_DATA, SOURCE_DATA):
        (directory / filename).write_text(payload, encoding="utf-8")


def number(value: object) -> int:
    if value in (None, ""):
        return 0
    return int(round(float(value)))


def rate(used: int, total: int) -> float:
    return used / total if total else 0.0


def load_territories():
    rows = json_load(PUBLIC_DATA / "municipios_index.json")
    if len(rows) != 158:
        raise ValueError(f"Expected 158 municipalities, found {len(rows)}")

    province_names = {}
    municipality_by_key = {}
    municipalities_by_province = defaultdict(list)
    for row in rows:
        province_key = normalize_name(row["provincia"])
        municipality_key = normalize_name(row["municipio"])
        province_names[province_key] = row["provincia"]
        municipality_by_key[(province_key, municipality_key)] = row
        municipalities_by_province[province_key].append(row)
    return rows, province_names, municipality_by_key, municipalities_by_province


def find_municipality(municipality_by_key, province_name: object, municipality_name: object):
    province_key = normalize_name(province_name)
    municipality_key = normalize_name(municipality_name)
    row = municipality_by_key.get((province_key, municipality_key))
    if row is None:
        raise KeyError(f"Unmapped municipality: {province_name!r} / {municipality_name!r}")
    return row


def level_values(ws, start_row: int, expected_indent: int):
    result = {}
    for row_index in range(start_row + 1, start_row + 6):
        cell = ws.cell(row_index, 1)
        level_key = LEVEL_KEYS.get(normalize_name(cell.value))
        if level_key is None or int(cell.alignment.indent or 0) != expected_indent:
            raise ValueError(
                f"Unexpected education-level row at {ws.title}!A{row_index}: "
                f"{cell.value!r}, indent={cell.alignment.indent!r}"
            )
        values = [number(ws.cell(row_index, column).value) for column in range(2, 11)]
        result[level_key] = {
            "total": values[0],
            "h": values[1],
            "m": values[2],
            "urbano_total": values[3],
            "urbano_h": values[4],
            "urbano_m": values[5],
            "rural_total": values[6],
            "rural_h": values[7],
            "rural_m": values[8],
        }
    return result


def rebuild_education_levels(index_rows, province_names, municipality_by_key):
    workbook = load_workbook(CENSUS_SOURCE / "cuadro-1-del-volumen-iv.xlsx", data_only=True)
    ws = workbook["C1"]

    municipalities = []
    provinces = []
    current_province = None

    for row_index in range(1, ws.max_row + 1):
        cell = ws.cell(row_index, 1)
        label = str(cell.value or "").strip()
        indent = int(cell.alignment.indent or 0)
        bold = bool(cell.font.bold)

        if not label:
            continue

        if indent == 0 and bold:
            current_province = None
            continue

        if indent == 1 and bold:
            province_key = normalize_name(label)
            if province_key not in province_names:
                raise KeyError(f"Unmapped education province: {label!r}")
            current_province = province_names[province_key]
            levels = level_values(ws, row_index, 5)
            provinces.append({"provincia": current_province, "nivel": levels})

            if province_key == "DISTRITO NACIONAL":
                territory = find_municipality(
                    municipality_by_key,
                    current_province,
                    "Santo Domingo de Guzmán",
                )
                municipalities.append({**territory, "nivel": levels})
            continue

        if indent != 3 or bold or not current_province:
            continue

        lookup_key = (normalize_name(current_province), normalize_name(label))
        territory = municipality_by_key.get(lookup_key)
        if territory is None:
            continue
        municipalities.append({**territory, "nivel": level_values(ws, row_index, 7)})

    municipality_codes = [row["adm2_code"] for row in municipalities]
    if len(municipalities) != 158 or len(set(municipality_codes)) != 158:
        duplicates = [code for code, count in Counter(municipality_codes).items() if count > 1]
        missing = sorted({row["adm2_code"] for row in index_rows} - set(municipality_codes))
        raise ValueError(
            f"Education level mapping failed: rows={len(municipalities)}, "
            f"duplicates={duplicates}, missing={missing}"
        )
    if len(provinces) != 32:
        raise ValueError(f"Expected 32 education provinces, found {len(provinces)}")

    national_levels = level_values(ws, 6, 3)
    national_total = number(ws.cell(6, 2).value)
    if sum(level["total"] for level in national_levels.values()) != national_total:
        raise ValueError("National education-level categories do not sum to their source total")

    national = {
        "poblacion_3_mas": national_total,
        "niveles": {
            key: {
                **values,
                "porcentaje": round(values["total"] / national_total * 100, 2),
            }
            for key, values in national_levels.items()
        },
    }

    municipalities.sort(key=lambda row: row["adm2_code"])
    provinces.sort(key=lambda row: normalize_name(row["provincia"]))
    write_asset("educacion_nivel.json", municipalities)
    write_asset("educacion_nivel_provincia.json", provinces)
    write_asset("national_educacion_nivel.json", national)


def parse_internet_rows(province_names, municipality_by_key):
    workbook = load_workbook(CENSUS_SOURCE / "cuadro-5-vol-4.xlsx", data_only=True)
    ws = workbook["C5"]
    municipality_values = {}
    province_values = {}
    current_province = None

    for row_index in range(1, ws.max_row + 1):
        cell = ws.cell(row_index, 1)
        label = str(cell.value or "").strip()
        indent = int(cell.alignment.indent or 0)
        bold = bool(cell.font.bold)

        if not label:
            continue

        if indent == 0 and bold:
            current_province = None
            continue
        if indent == 1 and bold:
            province_key = normalize_name(label)
            if province_key not in province_names:
                raise KeyError(f"Unmapped TIC province: {label!r}")
            current_province = province_names[province_key]
            total = number(ws.cell(row_index, 2).value)
            used = number(ws.cell(row_index + 1, 2).value)
            province_values[normalize_name(current_province)] = (total, used)
            if province_key == "DISTRITO NACIONAL":
                territory = find_municipality(
                    municipality_by_key,
                    current_province,
                    "Santo Domingo de Guzmán",
                )
                municipality_values[territory["adm2_code"]] = (total, used)
            continue
        if indent != 3 or bold or not current_province:
            continue
        territory = municipality_by_key.get(
            (normalize_name(current_province), normalize_name(label))
        )
        if territory is None:
            continue
        municipality_values[territory["adm2_code"]] = (
            number(ws.cell(row_index, 2).value),
            number(ws.cell(row_index + 1, 2).value),
        )

    national = (number(ws.cell(5, 2).value), number(ws.cell(6, 2).value))
    return municipality_values, province_values, national


def parse_device_rows(province_names, municipality_by_key):
    workbook = load_workbook(CENSUS_SOURCE / "cuadro-6-vol-4.xlsx", data_only=True)
    ws = workbook["C6"]
    municipality_values = {}
    province_values = {}
    current_province = None

    def values(row_index: int):
        total = number(ws.cell(row_index, 2).value)
        computer = number(ws.cell(row_index, 3).value) + number(ws.cell(row_index, 4).value)
        smartphone = number(ws.cell(row_index, 6).value)
        return total, computer, smartphone

    for row_index in range(1, ws.max_row + 1):
        cell = ws.cell(row_index, 1)
        label = str(cell.value or "").strip()
        indent = int(cell.alignment.indent or 0)
        bold = bool(cell.font.bold)

        if not label:
            continue

        if indent == 0 and bold:
            current_province = None
            continue
        if indent == 1 and bold:
            province_key = normalize_name(label)
            if province_key not in province_names:
                raise KeyError(f"Unmapped device province: {label!r}")
            current_province = province_names[province_key]
            province_values[province_key] = values(row_index)
            if province_key == "DISTRITO NACIONAL":
                territory = find_municipality(
                    municipality_by_key,
                    current_province,
                    "Santo Domingo de Guzmán",
                )
                municipality_values[territory["adm2_code"]] = values(row_index)
            continue
        if indent != 3 or bold or not current_province:
            continue
        territory = municipality_by_key.get(
            (normalize_name(current_province), normalize_name(label))
        )
        if territory is None:
            continue
        municipality_values[territory["adm2_code"]] = values(row_index)

    return municipality_values, province_values, values(5)


def tic_metric(total: int, used: int, municipality: str | None = None):
    value = {"total": total, "used": used, "rate_used": rate(used, total)}
    if municipality is not None:
        return {"municipio": municipality, **value}
    return value


def rebuild_tic(index_rows, province_names, municipality_by_key):
    internet_m, internet_p, internet_n = parse_internet_rows(
        province_names, municipality_by_key
    )
    devices_m, devices_p, devices_n = parse_device_rows(
        province_names, municipality_by_key
    )

    expected_codes = {row["adm2_code"] for row in index_rows}
    if set(internet_m) != expected_codes or set(devices_m) != expected_codes:
        raise ValueError(
            "TIC mapping incomplete: "
            f"internet missing={sorted(expected_codes - set(internet_m))}, "
            f"devices missing={sorted(expected_codes - set(devices_m))}"
        )

    municipalities = []
    for territory in sorted(index_rows, key=lambda row: row["adm2_code"]):
        code = territory["adm2_code"]
        internet_total, internet_used = internet_m[code]
        device_total, computer_used, smartphone_used = devices_m[code]
        if internet_total != device_total:
            raise ValueError(f"TIC population denominator mismatch for {code}")
        name = territory["municipio"]
        municipalities.append(
            {
                "adm2_code": code,
                "municipio": name,
                "internet": tic_metric(internet_total, internet_used, name),
                "cellular": tic_metric(device_total, smartphone_used, name),
                "computer": tic_metric(device_total, computer_used, name),
            }
        )

    provinces = []
    for province_key, province_name in province_names.items():
        internet_total, internet_used = internet_p[province_key]
        device_total, computer_used, smartphone_used = devices_p[province_key]
        if internet_total != device_total:
            raise ValueError(f"TIC population denominator mismatch for {province_name}")
        provinces.append(
            {
                "provincia": province_name,
                "internet": tic_metric(internet_total, internet_used),
                "cellular": tic_metric(device_total, smartphone_used),
                "computer": tic_metric(device_total, computer_used),
            }
        )

    internet_total, internet_used = internet_n
    device_total, computer_used, smartphone_used = devices_n
    if internet_total != device_total:
        raise ValueError("National TIC population denominator mismatch")
    national = {
        "internet": tic_metric(internet_total, internet_used),
        "cellular": tic_metric(device_total, smartphone_used),
        "computer": tic_metric(device_total, computer_used),
    }

    provinces.sort(key=lambda row: normalize_name(row["provincia"]))
    write_asset("tic.json", municipalities)
    write_asset("tic_provincia.json", provinces)
    write_asset("national_tic.json", national)


def rebuild_household_national():
    municipality_rows = json_load(PUBLIC_DATA / "hogares_resumen.json")
    households = sum(number(row.get("hogares_total")) for row in municipality_rows)
    household_population = sum(
        number(row.get("poblacion_en_hogares")) for row in municipality_rows
    )
    if households != 3_726_048:
        raise ValueError(f"Unexpected municipality household total: {households}")
    national = {
        "hogares_total": households,
        "poblacion_en_hogares": household_population,
        "personas_por_hogar": household_population / households,
    }
    write_asset("national_hogares.json", national)


def rebuild_population_2022(index_rows):
    if not FINAL_POPULATION_SOURCE.exists():
        raise FileNotFoundError(
            "Final population source is missing; run "
            "`node scripts/fetch_final_population_2022.mjs` first"
        )
    source = json_load(FINAL_POPULATION_SOURCE)
    source_rows = {
        str(row["adm2_code"]).zfill(5): row for row in source["municipalities"]
    }
    expected_codes = {row["adm2_code"] for row in index_rows}
    if set(source_rows) != expected_codes:
        raise ValueError(
            "Final population mapping differs from the municipality index: "
            f"missing={sorted(expected_codes - set(source_rows))}, "
            f"extra={sorted(set(source_rows) - expected_codes)}"
        )

    indicators = json_load(PUBLIC_DATA / "indicadores_basicos.json")
    for row in indicators:
        final = source_rows[str(row["adm2_code"]).zfill(5)]
        row["poblacion_total"] = number(final["poblacion_total"])
        row["poblacion_hombres"] = number(final["poblacion_hombres"])
        row["poblacion_mujeres"] = number(final["poblacion_mujeres"])
        population_2010 = row.get("poblacion_2010")
        if population_2010:
            row["variacion_abs"] = row["poblacion_total"] - number(population_2010)
            row["variacion_pct"] = round(
                row["variacion_abs"] / number(population_2010) * 100, 2
            )
        else:
            row["variacion_abs"] = None
            row["variacion_pct"] = None

    national_total = sum(row["poblacion_total"] for row in indicators)
    national_male = sum(row["poblacion_hombres"] for row in indicators)
    national_female = sum(row["poblacion_mujeres"] for row in indicators)
    if national_total != 10_773_983 or national_male + national_female != national_total:
        raise ValueError(
            "Final population national controls failed: "
            f"total={national_total}, male={national_male}, female={national_female}"
        )

    national = json_load(PUBLIC_DATA / "national_basic.json")
    national["poblacion_total"] = national_total
    national["poblacion_hombres"] = national_male
    national["poblacion_mujeres"] = national_female
    national["poblacion_2010"] = number(national.get("poblacion_total_2010"))
    write_asset("indicadores_basicos.json", indicators)
    write_asset("national_basic.json", national)


def rebuild_education_offer_national(province_names):
    municipalities = json_load(PUBLIC_DATA / "educacion_oferta_municipal.json")
    levels = {
        "inicial_primario": {"centros": 0, "matricula": 0},
        "secundario": {"centros": 0, "matricula": 0},
        "adultos": {"centros": 0, "matricula": 0},
    }
    province_aggregate = {
        key: {
            "provincia": name,
            "centros_total": 0,
            "niveles": {
                level: {"centros": 0, "matricula": 0} for level in levels
            },
        }
        for key, name in province_names.items()
    }

    total = 0
    for row in municipalities:
        province_key = normalize_name(row["provincia"])
        target = province_aggregate[province_key]
        row_levels = row.get("niveles") or {}
        row_total = number(row.get("centros_total"))
        total += row_total
        target["centros_total"] += row_total
        for level in levels:
            for field in ("centros", "matricula"):
                value = number((row_levels.get(level) or {}).get(field))
                levels[level][field] += value
                target["niveles"][level][field] += value

    if total != sum(value["centros"] for value in levels.values()):
        raise ValueError("Education-offer total does not equal its level totals")
    national = {"centros_total": total, "niveles": levels}
    provinces = sorted(
        province_aggregate.values(), key=lambda row: normalize_name(row["provincia"])
    )
    write_asset("educacion_oferta_municipal_provincia.json", provinces)
    write_asset("national_educacion_oferta.json", national)


def optional_float(value: object):
    text = str(value or "").strip().replace(",", ".")
    try:
        result = float(text)
    except ValueError:
        return None
    return result if math.isfinite(result) else None


def optional_int(value: object):
    text = str(value or "").strip()
    return int(text) if text.isdigit() else None


def rebuild_health(index_rows, province_names, municipality_by_key):
    by_municipality = {
        row["adm2_code"]: {
            "municipio": row["municipio"],
            "provincia": row["provincia"],
            "centros": [],
        }
        for row in index_rows
    }
    type_counts = Counter()
    region_counts = Counter()

    with HEALTH_SOURCE.open("r", encoding="utf-8-sig", newline="") as handle:
        source_rows = list(csv.DictReader(handle))

    unresolved = Counter()
    for source in source_rows:
        try:
            territory = find_municipality(
                municipality_by_key, source["PROVINCIA"], source["MUNICIPIO"]
            )
        except KeyError:
            unresolved[(source["PROVINCIA"], source["MUNICIPIO"])] += 1
            continue
        center_type = str(source["TIPO DE CENTRO"] or "").strip()
        health_region = str(source["REGIONAL DE SALUD"] or "").strip()
        center = {
            "adm2_code": territory["adm2_code"],
            "id_centro": optional_int(source["ID DEL CENTRO"]),
            "nombre": str(source["NOMBRE DEL ESTABLECIMIENTO"] or "").strip(),
            "tipo_centro": center_type,
            "regional_salud": health_region,
            "anio_apertura": optional_int(source["AÑO DE APERTURA"]),
            "latitud": optional_float(source["COORD. LATITUD"]),
            "longitud": optional_float(source["COORD. LONGITUD"]),
        }
        by_municipality[territory["adm2_code"]]["centros"].append(center)
        type_counts[center_type] += 1
        region_counts[health_region] += 1

    if unresolved:
        details = ", ".join(
            f"{province}/{municipality} ({count})"
            for (province, municipality), count in sorted(unresolved.items())
        )
        raise KeyError(f"Unmapped health territories: {details}")

    for row in by_municipality.values():
        row["centros"].sort(key=lambda center: (center["anio_apertura"] or 9999, center["id_centro"] or 0))

    province_centers = {key: [] for key in province_names}
    for row in by_municipality.values():
        province_centers[normalize_name(row["provincia"])].extend(row["centros"])
    provinces = [
        {"provincia": province_names[key], "centros": province_centers[key]}
        for key in province_names
    ]
    provinces.sort(key=lambda row: normalize_name(row["provincia"]))

    mapped_total = sum(len(row["centros"]) for row in by_municipality.values())
    if mapped_total != len(source_rows):
        raise ValueError(f"Health mapping lost rows: source={len(source_rows)}, mapped={mapped_total}")
    national = {
        "total_centros": mapped_total,
        "por_tipo": dict(sorted(type_counts.items())),
        "por_regional_salud": dict(sorted(region_counts.items())),
    }

    write_asset("salud_establecimientos.json", by_municipality)
    write_asset("salud_establecimientos_provincia.json", provinces)
    write_asset("national_salud_establecimientos.json", national)


def main():
    index_rows, province_names, municipality_by_key, _ = load_territories()
    rebuild_population_2022(index_rows)
    rebuild_education_levels(index_rows, province_names, municipality_by_key)
    rebuild_tic(index_rows, province_names, municipality_by_key)
    rebuild_household_national()
    rebuild_education_offer_national(province_names)
    rebuild_health(index_rows, province_names, municipality_by_key)
    print(
        "Rebuilt audited population, education, TIC, household, "
        "education-offer and health assets."
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        raise
